from flask import Flask, render_template, request, redirect, url_for, send_file
import os
from openpyxl import load_workbook

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads/'
app.config['OUTPUT_FOLDER'] = 'output/'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    if 'source_file' not in request.files or 'target_file' not in request.files:
        return "Both files are required."

    source_file = request.files['source_file']
    target_file = request.files['target_file']

    source_filepath = os.path.join(app.config['UPLOAD_FOLDER'], source_file.filename)
    target_filepath = os.path.join(app.config['UPLOAD_FOLDER'], target_file.filename)

    source_file.save(source_filepath)
    target_file.save(target_filepath)

    return redirect(url_for('process_data', source_file=source_file.filename, target_file=target_file.filename))

@app.route('/process/<source_file>/<target_file>')
def process_data(source_file, target_file):
    source_file_path = os.path.join(app.config['UPLOAD_FOLDER'], source_file)
    target_file_path = os.path.join(app.config['UPLOAD_FOLDER'], target_file)
    output_file_path = os.path.join(app.config['OUTPUT_FOLDER'], 'processed_' + target_file)

    # Load the source and target workbooks
    source_workbook = load_workbook(source_file_path, data_only=True)
    target_workbook = load_workbook(target_file_path, data_only=True)

    # Ensure "Financial Statements" sheet exists in the target workbook
    if "Financial Statements" not in target_workbook.sheetnames:
        return "Sheet 'Financial Statements' not found in the target file."

    target_sheet = target_workbook["Financial Statements"]

    # Define columns
    year_col = 'F'  # Target column for source column G data
    data_col = 'G'  # Target column for source column H data
    tab_col = 'N'   # Column containing the source sheet name
    row_col = 'M'   # Column containing the source row number

    # Starting point for values in source columns G and H
    source_start_row = 20

    # Iterate through rows in the target sheet, starting from row 6
    for target_row in range(6, target_sheet.max_row + 1):
        tab_value = target_sheet[f"{tab_col}{target_row}"].value  # Source sheet name
        source_row_value = target_sheet[f"{row_col}{target_row}"].value  # Source row number

        if tab_value and source_row_value:
            try:
                # Access the source sheet
                source_sheet = source_workbook[tab_value.strip()]

                # Calculate the actual source row
                source_row = source_start_row + int(source_row_value) - 1

                # Extract values from source columns G and H
                if source_row <= source_sheet.max_row:
                    value_g = source_sheet.cell(row=source_row, column=7).value  # Column G
                    value_h = source_sheet.cell(row=source_row, column=8).value  # Column H
                else:
                    value_g = None
                    value_h = None

                # Write values to the correct row in the target sheet, aligning to row 4
                corrected_target_row = target_row - 2  # Adjust target_row to align data starting from row 4
                target_sheet[f"{year_col}{corrected_target_row}"] = value_g
                target_sheet[f"{data_col}{corrected_target_row}"] = value_h

                print(f"Updated Target Row {corrected_target_row}: F{corrected_target_row} with '{value_g}', G{corrected_target_row} with '{value_h}' from Source Row {source_row}")

            except KeyError:
                print(f"Sheet '{tab_value}' not found in source file.")
            except ValueError:
                print(f"Invalid source row value '{source_row_value}' in row {target_row}.")
            except Exception as e:
                print(f"Error processing target row {target_row}: {e}")
        else:
            print(f"Skipping target row {target_row} due to missing Tab or Row value.")

    # Save the updated target workbook
    target_workbook.save(output_file_path)
    print(f"Updated target sheet saved as '{output_file_path}'")

    # Extract a preview starting from row 3 to row 243
    preview_data = []
    for row in target_sheet.iter_rows(min_row=3, max_row=243, values_only=True):
        cleaned_row = [cell if cell is not None else "" for cell in row]  # Replace None with empty string
        preview_data.append(cleaned_row)

    return render_template('preview.html', preview_data=preview_data, output_file='processed_' + target_file)


@app.route('/download/<filename>')
def download_file(filename):
    path = os.path.join(app.config['OUTPUT_FOLDER'], filename)
    return send_file(path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
